from azure.cognitiveservices.vision.customvision.prediction import CustomVisionPredictionClient
from matplotlib.backend_bases import FigureManagerBase #importing azure custom vision prediction client
from msrest.authentication import ApiKeyCredentials #for api key credentials from ms
from matplotlib import pyplot as plt #for plotting
from PIL import Image, ImageDraw, ImageFont #for image processing
import numpy as np # for array processing
import os #for file handling
import streamlit as st #for deployment
import datetime #for date and time
import openpyxl #for excel file handling
import smtplib
from email.mime.text import MIMEText
from email.header import Header

def log_attendance(student_name, date_time):
    excel_file = 'attendance.xlsx'
    if not os.path.isfile(excel_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Date/Time"])
    else:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    
    ws.append([student_name, date_time])
    wb.save(excel_file)


def main():
    from dotenv import load_dotenv #importing dotenv for environment variables and loading them
    
    try:
        load_dotenv()
        prediction_endpoint = os.getenv('PREDICTION_ENDPOINT')
        prediction_key = os.getenv('PREDICTION_KEY')
        project_id = os.getenv('PROJECT_ID')
        model_name = os.getenv('PUBLISHED_MODEL_NAME')
        email_sender = os.getenv('EMAIL_SENDER')
        email_password = os.getenv('EMAIL_PASSWORD')
        email_receiver = os.getenv('EMAIL_RECEIVER')

        credentials = ApiKeyCredentials(in_headers={"Prediction-key": prediction_key})
        predicton_client = CustomVisionPredictionClient(endpoint=prediction_endpoint, credentials=credentials)
    
        image_path = 'test_image.jpg' #path to the test image
        print('Loading image from: ' + image_path)
        image = Image.open(image_path)
        h, w, ch = np.array(image).shape
        print('Image size: ' + str(w) + 'x' + str(h)) #printing the size of the image

        with open(image_path, mode='rb') as image_data: 
            results = predicton_client.detect_image(project_id, model_name, image_data.read()) #detecting the image

        fig = plt.figure(figsize=(8, 8)) 
        plt.axis('off')
        plt.title('IntelliSense')
        plt.imshow(image)

        draw = ImageDraw.Draw(image)
        lineWidth = int(w/100)
        color = 'red'
        font = ImageFont.truetype('arial.ttf', size=int(w/50))

        #printing the predictions
        for prediction in results.predictions: 
            if prediction.probability > 0.7:
                x = prediction.bounding_box.left * w
                y = prediction.bounding_box.top * h
                width = prediction.bounding_box.width * w
                height = prediction.bounding_box.height * h
                points = ((x, y), (x + width, y), (x + width, y + height), (x , y + height), (x, y))
                draw.line(points, fill=color, width=lineWidth)
                draw.text((x, y), prediction.tag_name, fill=color, font=font)
                plt.annotate(prediction.tag_name + ': {0:.2f}%'.format(prediction.probability * 100), xy=(x, y), xytext=(x, y), color=color, fontsize=12, ha='left', va='bottom')
                print('\t' + prediction.tag_name + ': {0:.2f}%'.format(prediction.probability * 100)) #printing the accuracy of the prediction
                #confidence level
                if prediction.probability > 0.7:
                    print('High confidence level, you can trust the prediction result')
                else:
                    print('Low confidence level')

                #log the attendance in excel file
                student_name = prediction.tag_name
                date_time = datetime.datetime.now()
                log_attendance(student_name, date_time)

                excel_file = 'attendance.xlsx'
                if not os.path.isfile(excel_file):
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.append(["Name", "Date/Time"])
                else:
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
    
                    ws.append([student_name, date_time])
                    wb.save(excel_file)

                    #attendance alert function: if face recognized then send email

                    subject = 'Attendance Alert'
                    now = datetime.datetime.now()
                    date_time = now.strftime("%Y-%m-%d %H:%M:%S")
                    body = "Hi " + student_name + ",\n\nYour attendance has been marked at " + date_time + ".\n\nRegards,\nIntelliSense"

                    em = MIMEText(body, 'plain', 'utf-8')
                    em['From'] = Header(email_sender, 'utf-8')
                    em['To'] = Header(email_receiver, 'utf-8')
                    em['Subject'] = Header(subject, 'utf-8')

                    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                        smtp.login(email_sender, email_password)
                        smtp.send_message(em)
                        print('Email sent successfully to ' + student_name)

        #saving the output image and displaying it
        fig.canvas.manager.window.wm_iconbitmap('ai.ico')
        plt.show()
        output_path = os.path.join('test_image_output.jpg')
        fig.savefig(output_path)
        print('Saved output image to: ' + output_path)
        now = datetime.datetime.now()
        print("Attendance marked at: ", now.strftime("%Y-%m-%d %H:%M:%S"))
    except Exception as ex:
        print(ex)

if __name__ == '__main__':
    main()

#streamlit app
st.set_page_config(layout='wide')
st.title("IntelliSense")
st.write("Smart Surveillance and Attendance System for Classrooms")

input_image = st.file_uploader("Upload an image", type=["jpg", "jpeg", "png"])

if input_image is not None:
    if st.button("Predict"):
        
        col1, col2, col3 = st.columns([1,1,1])

        with col1:
            st.info("Your uploaded Image")
            image = Image.open(input_image)
            st.image(image, use_column_width=True)

        with col2:
            st.info("Output Image with Prediction")
            main()
            st.image("test_image_output.jpg", use_column_width=True)
            st.balloons()
            

st.sidebar.title("About")
st.sidebar.info("This is a smart surveillance and attendance system for classrooms.")

st.sidebar.title("How it works?")
st.sidebar.info("Upload an image of students and click on predict. The model will detect the students and mark their attendance.")

st.sidebar.title("Author")
st.sidebar.info("Krishna Agarwal")

st.sidebar.title("Built with")
st.sidebar.info("This web app is built using Streamlit, Azure Custom Vision and Python.")
