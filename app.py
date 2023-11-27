from azure.cognitiveservices.vision.customvision.prediction import CustomVisionPredictionClient #importing azure custom vision prediction client
from msrest.authentication import ApiKeyCredentials #for api key credentials from ms
from matplotlib import pyplot as plt #for plotting
from PIL import Image, ImageDraw, ImageFont #for image processing
import numpy as np # for array processing
import os #for file handling
import streamlit as st #for deployment
import datetime #for date and time
import openpyxl #for excel file handling
from openpyxl.utils import get_column_letter #for excel file handling
import smtplib #for emailing
from email.mime.text import MIMEText 
from email.header import Header
import cv2 #for webcam input
import time #for time 
import pandas as pd #for excel file handling


def main():
    from dotenv import load_dotenv #importing dotenv for environment variables and loading them
    
    try:
        load_dotenv()
        prediction_endpoint = os.getenv('PREDICTION_ENDPOINT')
        prediction_key = os.getenv('PREDICTION_KEY')
        project_id = os.getenv('PROJECT_ID')
        model_name = os.getenv('PUBLISHED_MODEL_NAME')
        email_sender = os.getenv('EMAIL_SENDER')
        email_password = os.getenv('EMAIL_PASSWORD')
        email_receiver = os.getenv('EMAIL_RECEIVER')
        smtp_server = os.getenv('SMTP_SERVER')
        smtp_port = os.getenv('SMTP_PORT')

        credentials = ApiKeyCredentials(in_headers={"Prediction-key": prediction_key})
        predicton_client = CustomVisionPredictionClient(endpoint=prediction_endpoint, credentials=credentials)

        print("Welcome! The camera is warming up and will be ready in a moment. Please be patient, this may take a minute or two.")
        cv2.namedWindow("IntelliSense")
        cap = cv2.VideoCapture(1) #can change the camera number if multiple cameras are connected
        
        #printing a message to the user that the camera is initializing and will be ready in a moment

        time.sleep(3)
        
        attendance_data = []

        while True:
            ret, frame = cap.read() 
            cv2.imshow("IntelliSense", frame) 
            if not ret: 
                break 
            cv2.waitKey(2) 

            time.sleep(5) 

            img_name = "test_image.jpg" 
            cv2.imwrite(img_name, frame)
            print("{} written!".format(img_name))
            break

        cap.release()
        cv2.destroyAllWindows()

        image_path = 'test_image.jpg' #path to the test image file
        print('Loading image from: ' + image_path)
        image = Image.open(image_path)
        h, w, ch = np.array(image).shape
        print('Image size: ' + str(w) + 'x' + str(h)) #printing the size of the image

        with open(image_path, mode='rb') as image_data: 
            results = predicton_client.detect_image(project_id, model_name, image_data.read()) #detecting the image


        fig = plt.figure(figsize=(8, 8))
        plt.axis('on')
        color = 'red'

        draw = ImageDraw.Draw(image)
        lineWidth = int(w/100)
        font = ImageFont.truetype('arial.ttf', size=int(w/50)) 

        # Get the red value of the pixel at the center of the image
        red_value = np.array(image)[h//2, w//2, 2]
        red_normalized = red_value / 255.0
        # Get the green value of the pixel at the center of the image
        green_value = np.array(image)[h//2, w//2, 1]
        green_normalized = green_value / 255.0
        # Get the blue value of the pixel at the center of the image
        blue_value = np.array(image)[h//2, w//2, 0]
        blue_normalized = blue_value / 255.0


        #printing the predictions
        for prediction in results.predictions:
            if (prediction.probability*100) > 75:
                left = prediction.bounding_box.left * w
                top = prediction.bounding_box.top * h
                width = prediction.bounding_box.width * w
                height = prediction.bounding_box.height * h

                # Draw the bounding box with red color, along with the predicted class name and probability in percentage
                points = ((left, top), (left + width, top), (left + width, top + height), (left, top + height), (left, top))
                draw.line(points, fill=color, width=lineWidth)
                plt.annotate(prediction.tag_name + ": {0:.2f}%".format(prediction.probability * 100),(left,top), backgroundcolor=color)
                #draw.rectangle(((x, y), (x + width, y + height)), outline=box_color, width=lineWidth)
                #draw.text((x, y), prediction.tag_name + ': ' + str(round(prediction.probability * 100, 2)) + '%', fill=box_color, font=font)                
                print(f"\t{prediction.tag_name}: {prediction.probability * 100:.2f}%")      
                plt.title('IntelliSense')
                plt.imshow(image)
                plt.pause(5)
                
                #confidence level
                if prediction.probability > 0.7:
                    print('High confidence level, you can trust the prediction result')
                else:
                    print('Low confidence level')


                #Logging the Attendance Data in Excel File - Start

                student_name = prediction.tag_name
                date_time = datetime.datetime.now()

                #Here, we defined a function to find the next available row in the excel file, starting from row 14.
                def find_next_row(sheet):
                    current_row = 14
                    while sheet[get_column_letter(1) + str(current_row)].value is not None:
                        current_row += 1
                    return current_row
                
                #Here, we defined a function to write the data to the specified columns.
                def write_recognition_data(file_path, date, timestamp, student_recognized):

                    # Load the workbook
                    workbook = openpyxl.load_workbook(file_path)
                
                    # Select the active sheet
                    sheet = workbook.active #sheet that is currently being used, can specify the sheet name
                
                    # Find the next available row dynamically
                    next_row = find_next_row(sheet) #calling the function defined above
                
                    # Write data to the specified columns
                    sheet[get_column_letter(1) + str(next_row)] = date.strftime("%d %B %Y")
                    sheet[get_column_letter(2) + str(next_row)] = timestamp.strftime("%I:%M:%S %p")
                    sheet[get_column_letter(3) + str(next_row)] = student_recognized
                
                    # Save the changes
                    workbook.save(file_path)

                # Setting the variables for the function
                file_path = "attendance.xlsx"
                date = datetime.date.today()
                timestamp = datetime.datetime.now()
                student_recognized = student_name

                # Calling the function that writes the data to the excel file.
                write_recognition_data(file_path, date, timestamp, student_recognized)
                
                # Logging the Attendance Data in Excel File - End


                # Emailing The Student - Attendance Alert - Start

                #retrieving the email id of the student recognized

                #Specify the path of the excel file
                excel_file = 'attendance.xlsx'
                #Load the excel file into a pandas dataframe, skipping the header and first 3 rows
                df = pd.read_excel(excel_file, header=None, skiprows=3)
                #Define the column names based on your previous example
                column_names = ['Student ID', 'Name', 'DOB', 'Email']
                #Set the dataframe's column names to the ones you just defined
                df.columns = column_names
                #searching the student name in the excel file who is recognized by the model
                search_name = student_name
                #find the row in the dataframe where the name matches the search name
                search_results = df.loc[df['Name'] == search_name]
                #check if the search results are not empty
                if not search_results.empty:
                    email_id = search_results['Email'].iloc[0]
                    print(f"Student: {search_name}, Email: {email_id} found in the database.")
                else:
                    print(f"Student: {search_name} not found in the database.")
                    email_id = 'ceo.intellisense@gmail.com'

                #sending email to the student with attendance alert - mail body
                subject = 'Attendance Alert ' + 'for ' + student_name
                now = datetime.datetime.now() 
                date_time = now.strftime("%d %B %Y, %I:%M:%S %p") 

                body = "Dear " + student_name + ",\n\nWe want to inform you that your attendance has been successfully marked for the class on " + date_time + ".\nWe appreciate your dedication to your studies.\n\nRegards,\nIntelliSense\n\nThis is an automated email. Please do not reply to this email." 

                email_receiver = email_id
                email_message = MIMEText(body, 'plain', 'utf-8')
                email_message['From'] = Header(email_sender, 'utf-8')
                email_message['To'] = Header(email_receiver, 'utf-8')
                email_message['Subject'] = Header(subject, 'utf-8')

                #Send the message via our own SMTP server.
                with smtplib.SMTP_SSL(smtp_server, smtp_port) as smtp:
                    smtp.login(email_sender, email_password)
                    smtp.send_message(email_message)
                    print('Email sent successfully to ' + student_name + ' at ' + email_id)

                # Emailing The Student - Attendance Alert - End

        #Saves the output image with the bounding boxes and prediction results

        output_path = os.path.join('test_image_output.jpg')
        fig.savefig(output_path)
        print('Saved output image to: ' + output_path)
        now = datetime.datetime.now()
        print("Attendance marked and Saved in Database at: ", now.strftime("%d %B %Y, %I:%M:%S %p"))
    except Exception as ex:
        print(f"An error occurred: {ex}")

if __name__ == '__main__':
    main()


#Streamlit Web App Code

st.set_page_config(layout='wide')
st.title("IntelliSense")
st.write("Smart Surveillance and Attendance System for Classrooms")

input_image = st.file_uploader("Upload an image", type=["jpg", "jpeg", "png"])

if input_image is not None:
    if st.button("Predict"):
        
        col1, col2, col3 = st.columns([1,1,1])

        with col1:
            st.info("Your uploaded Image")
            image = Image.open(input_image)
            st.image(image, use_column_width=True)

        with col2:
            st.info("Output Image with Prediction")
            main()
            st.image("test_image_output.jpg", use_column_width=True)
            st.balloons()
            
st.sidebar.title("About")
st.sidebar.info("This is a smart surveillance and attendance system for classrooms.")

st.sidebar.title("How it works?")
st.sidebar.info("Upload an image of students and click on predict. The model will detect the students and mark their attendance.")

st.sidebar.title("Author")
st.sidebar.info("Krishna Agarwal")

st.sidebar.title("Built with")
st.sidebar.info("This web app is built using Streamlit, Azure Custom Vision and Python.")
